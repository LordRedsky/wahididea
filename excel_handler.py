"""
Excel Handler Module
Manages reading and writing data to Rekap.xlsx
"""

import os
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl import load_workbook


class ExcelHandler:
    """Handle Excel operations for storing extracted data"""
    
    # Expected headers - complete list
    EXPECTED_HEADERS = [
        "No",
        "Tanggal Upload",
        "Nama Pasien",
        "Tanggal Pemeriksaan",
        "ID Pasien",
        "Umur Pasien",
        "Jenis Kelamin",
        "Jenis Pemeriksaan",
        "Tegangan (kV)",
        "CTDIvol (mGy)",
        "Total DLP (mGy·cm)"
    ]
    
    def __init__(self, filename: str = "Rekap.xlsx"):
        self.filename = filename
        self._initialize_excel()
    
    def _initialize_excel(self):
        """Initialize Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.filename):
            self._create_new_excel()
        else:
            # Check and update headers if needed
            self._update_headers_if_needed()
    
    def _create_new_excel(self):
        """Create new Excel file with headers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Pasien"

        # Headers - Complete list of all fields
        headers = self.EXPECTED_HEADERS

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Style headers
        self._style_headers(ws)

        wb.save(self.filename)
    
    def _update_headers_if_needed(self):
        """Update Excel headers if file exists with old headers"""
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            
            # Get current headers
            current_headers = [cell.value for cell in ws[1]]
            
            # Check if headers match expected
            if current_headers != self.EXPECTED_HEADERS:
                print(f"Updating headers...")
                print(f"Old headers: {current_headers}")
                print(f"New headers: {self.EXPECTED_HEADERS}")
                
                # Update header row
                for col, header in enumerate(self.EXPECTED_HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Style headers
                self._style_headers(ws)
                
                wb.save(self.filename)
                print("Headers updated successfully!")
        except Exception as e:
            print(f"Error updating headers: {e}")
    
    def _style_headers(self, ws):
        """Apply styling to header row"""
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Auto-adjust column widths
        column_widths = [8, 15, 25, 20, 15, 12, 12, 30, 10, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def add_record(self, data: Dict[str, Optional[str]]) -> int:
        """
        Add a new record to Excel file

        Args:
            data: Dictionary with extracted patient data

        Returns:
            Row number where data was added
        """
        wb = load_workbook(self.filename)
        ws = wb.active

        # Get next row number
        next_row = ws.max_row + 1

        # Prepare data - match all fields
        record = [
            next_row - 1,  # No (sequential)
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Tanggal Upload
            data.get('nama_pasien', ''),
            data.get('tanggal_pemeriksaan', ''),
            data.get('id_pasien', ''),
            data.get('umur_pasien', ''),
            data.get('jenis_kelamin', ''),
            data.get('jenis_pemeriksaan', ''),
            data.get('kv', ''),
            data.get('ctdi_vol', ''),
            data.get('total_dlp', '')
        ]

        # Write data
        for col, value in enumerate(record, 1):
            ws.cell(row=next_row, column=col, value=value)

        wb.save(self.filename)
        return next_row - 1
    
    def add_records(self, records: List[Dict[str, Optional[str]]]) -> int:
        """
        Add multiple records to Excel file
        
        Args:
            records: List of dictionaries with patient data
            
        Returns:
            Number of records added
        """
        count = 0
        for record in records:
            self.add_record(record)
            count += 1
        return count
    
    def get_all_records(self) -> List[Dict]:
        """Get all records from Excel file"""
        if not os.path.exists(self.filename):
            return []
        
        wb = load_workbook(self.filename)
        ws = wb.active
        
        records = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            records.append(record)
        
        return records
    
    def get_record_count(self) -> int:
        """Get total number of records"""
        if not os.path.exists(self.filename):
            return 0
        
        wb = load_workbook(self.filename)
        ws = wb.active
        return ws.max_row - 1  # Subtract header row
    
    def clear_all_data(self):
        """Clear all data (keep headers)"""
        if not os.path.exists(self.filename):
            return
        
        wb = load_workbook(self.filename)
        ws = wb.active
        
        # Delete all rows except header
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)

        wb.save(self.filename)
